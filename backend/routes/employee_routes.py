from fastapi import APIRouter, HTTPException, status, Depends
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorDatabase
from pydantic import BaseModel
from typing import List, Optional
from models.user import UserRole
from models.verification import VerificationStatus
from middleware.auth_middleware import get_current_user
from services.audit_service import write_audit_log
from datetime import datetime, date, timedelta, timezone
import logging
import io
import csv
import asyncio
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/employee", tags=["Employee"])


class RMReviewRequest(BaseModel):
    remarks: Optional[str] = None


class RMRejectRequest(BaseModel):
    reason: str

async def require_employee(current_user: dict = Depends(get_current_user)):
    """Dependency to check if user is employee."""
    if current_user["role"] != UserRole.EMPLOYEE.value:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Employee access required"
        )
    return current_user

async def get_db():
    from server import db_instance
    return db_instance


def _clean_identifier(value):
    text = str(value or "").strip()
    return text if text else None


async def _get_rm_identifiers(db: AsyncIOMotorDatabase, current_user_or_id):
    if isinstance(current_user_or_id, dict):
        rm_user = current_user_or_id
        rm_id = rm_user.get("user_id")
    else:
        rm_id = current_user_or_id
        rm_user = await db.users.find_one({"user_id": rm_id, "role": "employee"}, {"_id": 0}) or {}

    identifiers = {
        _clean_identifier(rm_id),
        _clean_identifier(rm_user.get("user_id")),
        _clean_identifier(rm_user.get("employee_code")),
        _clean_identifier(rm_user.get("uid")),
        _clean_identifier(rm_user.get("rm_id")),
    }
    return [identifier for identifier in identifiers if identifier]


async def _get_employee_profile(db: AsyncIOMotorDatabase, current_user_or_id):
    if isinstance(current_user_or_id, dict):
        user_id = current_user_or_id.get("user_id")
        fallback = current_user_or_id
    else:
        user_id = current_user_or_id
        fallback = {}
    profile = await db.users.find_one({"user_id": user_id, "role": "employee"}, {"_id": 0})
    return profile or fallback or {}


def _is_branch_manager_profile(profile: dict) -> bool:
    role_key = str(profile.get("admin_role_key") or profile.get("role_key") or "").strip().lower()
    designation = str(profile.get("designation") or "").strip().lower()
    return role_key == "branch_manager" or "branch manager" in designation


async def _get_branch_manager_rm_users(db: AsyncIOMotorDatabase, branch_manager_profile: dict):
    branch_manager_identifiers = await _get_rm_identifiers(db, branch_manager_profile)
    rm_lookup_or = [
        *_field_matches_identifiers("reports_to", branch_manager_identifiers),
        *_field_matches_identifiers("branch_manager_id", branch_manager_identifiers),
    ]
    if not rm_lookup_or:
        return []
    return await db.users.find(
        {
            "role": "employee",
            "$or": rm_lookup_or,
            "$and": [{
                "$or": [
                    {"admin_role_key": {"$in": ["rm", "relationship_manager"]}},
                    {"designation": {"$regex": "relationship manager|\\brm\\b", "$options": "i"}},
                ]
            }],
        },
        {"_id": 0, "user_id": 1, "employee_code": 1, "uid": 1, "rm_id": 1, "full_name": 1, "is_active": 1}
    ).to_list(length=1000)


def _field_matches_identifiers(field, identifiers):
    exact_values = [value for value in identifiers if value]
    regex_values = [
        {"$regex": f"^{re.escape(value)}$", "$options": "i"}
        for value in identifiers
        if value
    ]
    return [{field: {"$in": exact_values}}] + [{field: regex} for regex in regex_values]


def _verification_scope_query(broker_ids, property_ids, identifiers):
    scope_or = _field_matches_identifiers("rm_id", identifiers)
    if broker_ids:
        scope_or.append({"broker_id": {"$in": broker_ids}})
    if property_ids:
        scope_or.append({"property_id": {"$in": property_ids}})
    return {"$or": scope_or}


def _empty_assignment_query(field):
    return {"$or": [{field: {"$exists": False}}, {field: None}, {field: ""}]}


def _assigned_ref(source: dict | None, field: str) -> str:
    value = str((source or {}).get(field) or "").strip()
    return "" if value in {"-", "NA", "N/A"} else value


async def _property_owner_assignment(db: AsyncIOMotorDatabase, property_data: dict) -> tuple[dict, dict]:
    owner = {}
    if property_data.get("owner_id"):
        owner = await db.users.find_one({"user_id": property_data["owner_id"], "role": "host"}, {"_id": 0}) or {}
    owner_found = bool(owner.get("user_id"))
    return owner, {
        "broker_id": _assigned_ref(owner, "broker_id") if owner_found else _assigned_ref(property_data, "broker_id"),
        "rm_id": _assigned_ref(owner, "rm_id") if owner_found else _assigned_ref(property_data, "rm_id"),
        "branch_manager_id": _assigned_ref(owner, "branch_manager_id") if owner_found else _assigned_ref(property_data, "branch_manager_id"),
        "branch_manager_code": _assigned_ref(owner, "branch_manager_code") if owner_found else _assigned_ref(property_data, "branch_manager_code"),
    }


def _pending_review_query(property_ids, identifiers, is_branch_manager):
    if is_branch_manager:
        return {
            "status": VerificationStatus.COMPLETED.value,
            "property_id": {"$in": property_ids},
            "branch_manager_reviewed": {"$ne": True},
            "$or": _field_matches_identifiers("branch_manager_id", identifiers),
        }
    return {
        "status": VerificationStatus.COMPLETED.value,
        "property_id": {"$in": property_ids},
        "rm_reviewed": False,
        "$and": [
            _empty_assignment_query("branch_manager_id"),
            {"broker_id": {"$nin": [None, ""]}},
        ],
        "$or": _field_matches_identifiers("rm_id", identifiers),
    }


def _history_review_query(property_ids, identifiers, is_branch_manager):
    if is_branch_manager:
        return {
            "property_id": {"$in": property_ids},
            "$or": _field_matches_identifiers("branch_manager_id", identifiers),
        }
    return {
        "property_id": {"$in": property_ids},
        "$or": _field_matches_identifiers("rm_id", identifiers),
    }


async def _get_rm_scope(db: AsyncIOMotorDatabase, current_user_or_id):
    """Return brokers and hosts visible to an RM across legacy and current assignment fields."""
    employee_profile = await _get_employee_profile(db, current_user_or_id)
    if _is_branch_manager_profile(employee_profile):
        branch_manager_identifiers = await _get_rm_identifiers(db, employee_profile)
        rm_users = await _get_branch_manager_rm_users(db, employee_profile)

        all_broker_ids = set()
        all_host_ids = set()
        for rm_user in rm_users:
            broker_ids, host_ids = await _get_rm_scope(db, rm_user)
            all_broker_ids.update(broker_ids)
            all_host_ids.update(host_ids)

        direct_bm_host_or = _field_matches_identifiers("branch_manager_id", branch_manager_identifiers)
        direct_bm_host_or.extend(_field_matches_identifiers("branch_manager_code", branch_manager_identifiers))
        bm_hosts = await db.users.find(
            {"role": "host", "$or": direct_bm_host_or},
            {"_id": 0, "user_id": 1, "broker_id": 1}
        ).to_list(length=3000)
        all_host_ids.update(host.get("user_id") for host in bm_hosts if host.get("user_id"))
        all_broker_ids.update(host.get("broker_id") for host in bm_hosts if host.get("broker_id"))

        return list(all_broker_ids), list(all_host_ids)

    identifiers = await _get_rm_identifiers(db, current_user_or_id)
    if not identifiers:
        return [], []

    direct_host_or = _field_matches_identifiers("rm_id", identifiers)
    direct_hosts = await db.users.find(
        {"role": "host", "$or": direct_host_or},
        {"_id": 0, "user_id": 1, "broker_id": 1}
    ).to_list(length=3000)
    direct_host_ids = list({host["user_id"] for host in direct_hosts if host.get("user_id")})

    direct_property_or = _field_matches_identifiers("rm_id", identifiers)
    if direct_host_ids:
        direct_property_or.append({"owner_id": {"$in": direct_host_ids}})
    direct_properties = await db.properties.find(
        {"$or": direct_property_or},
        {"_id": 0, "owner_id": 1, "broker_id": 1}
    ).to_list(length=5000)

    derived_broker_ids = {
        host.get("broker_id")
        for host in direct_hosts
        if host.get("broker_id")
    } | {
        prop.get("broker_id")
        for prop in direct_properties
        if prop.get("broker_id")
    }

    broker_or = _field_matches_identifiers("rm_id", identifiers)
    if derived_broker_ids:
        broker_or.append({"user_id": {"$in": list(derived_broker_ids)}})
    brokers = await db.users.find(
        {"role": "broker", "$or": broker_or},
        {"_id": 0, "user_id": 1}
    ).to_list(length=1000)
    broker_ids = list({broker["user_id"] for broker in brokers if broker.get("user_id")})

    host_or = [*direct_host_or]
    if broker_ids:
        host_or.append({"broker_id": {"$in": broker_ids}})
    property_owner_ids = [prop.get("owner_id") for prop in direct_properties if prop.get("owner_id")]
    if property_owner_ids:
        host_or.append({"user_id": {"$in": property_owner_ids}})
    hosts = await db.users.find(
        {"role": "host", "$or": host_or},
        {"_id": 0, "user_id": 1}
    ).to_list(length=3000)
    host_ids = list({host["user_id"] for host in hosts if host.get("user_id")})
    return broker_ids, host_ids

# ========== EMPLOYEE DASHBOARD ==========

@router.get("/dashboard/stats")
async def get_employee_dashboard_stats(
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Get employee dashboard statistics."""
    try:
        rm_id = current_user["user_id"]
        employee_profile = await _get_employee_profile(db, current_user)
        is_branch_manager = _is_branch_manager_profile(employee_profile)
        branch_manager_rms = await _get_branch_manager_rm_users(db, employee_profile) if is_branch_manager else []
        rm_identifiers = await _get_rm_identifiers(db, current_user)
        now = datetime.now(timezone.utc)
        today = date.today()
        month_start = today.replace(day=1)
        next_month = date(today.year + (1 if today.month == 12 else 0), 1 if today.month == 12 else today.month + 1, 1)

        broker_ids, my_host_ids = await _get_rm_scope(db, current_user)
        broker_cursor = db.users.find(
            {"user_id": {"$in": broker_ids}, "role": "broker"},
            {"_id": 0, "user_id": 1, "is_active": 1}
        )
        my_brokers = await broker_cursor.to_list(length=1000)
        my_broker_ids = [broker["user_id"] for broker in my_brokers]
        total_brokers = len(my_broker_ids)
        active_brokers = len([broker for broker in my_brokers if broker.get("is_active", True)])
        inactive_brokers = total_brokers - active_brokers

        host_cursor = db.users.find(
            {"role": "host", "user_id": {"$in": my_host_ids}},
            {"_id": 0, "user_id": 1, "kyc_status": 1}
        )
        my_hosts = await host_cursor.to_list(length=3000)
        my_host_ids = list({host["user_id"] for host in my_hosts if host.get("user_id")})
        pending_host_verification = len([host for host in my_hosts if host.get("kyc_status") not in {"approved", "verified"}])

        property_query = {"$or": [{"broker_id": {"$in": my_broker_ids}}, {"owner_id": {"$in": my_host_ids}}, *_field_matches_identifiers("rm_id", rm_identifiers), *_field_matches_identifiers("branch_manager_id", rm_identifiers)]}
        properties = await db.properties.find(
            property_query,
            {"_id": 0, "property_id": 1, "status": 1, "rating": 1}
        ).to_list(length=5000)
        property_ids = list({prop["property_id"] for prop in properties if prop.get("property_id")})

        live_properties = len([prop for prop in properties if prop.get("status") == "live"])
        pending_property_verification = len([prop for prop in properties if prop.get("status") in {"pending_verification", "under_review"}])
        rejected_properties = len([prop for prop in properties if prop.get("status") == "rejected"])
        draft_properties = len([prop for prop in properties if prop.get("status") == "draft"])
        
        review_properties = await db.properties.find(
            property_query,
            {"_id": 0, "property_id": 1}
        ).to_list(length=5000)
        review_property_ids = [p["property_id"] for p in review_properties if p.get("property_id")]
        
        # Pending verifications (all verifications waiting for this employee's review)
        pending_verifications = await db.property_verifications.count_documents(
            _pending_review_query(review_property_ids, rm_identifiers, is_branch_manager)
        )
        
        properties_under_review = await db.properties.count_documents({
            "status": "under_review",
            "property_id": {"$in": review_property_ids}
        })
        
        expiring_soon_date = (today + timedelta(days=5)).isoformat()

        expiring_subscriptions = await db.subscriptions.count_documents({
            "status": {"$in": ["trial", "active"]},
            "end_date": {"$lte": expiring_soon_date},
            "user_id": {"$in": my_host_ids}
        })

        booking_query = {
            "$or": [
                {"broker_id": {"$in": my_broker_ids}},
                {"user_broker_id": {"$in": my_broker_ids}},
                {"host_id": {"$in": my_host_ids}},
                {"owner_id": {"$in": my_host_ids}},
                {"user_id": {"$in": my_host_ids}},
                {"property_id": {"$in": property_ids}},
                *_field_matches_identifiers("rm_id", rm_identifiers),
                *_field_matches_identifiers("user_rm_id", rm_identifiers),
                *_field_matches_identifiers("employee_id", rm_identifiers),
                *_field_matches_identifiers("assigned_rm_id", rm_identifiers),
            ]
        }
        bookings = await db.bookings.find(
            booking_query,
            {"_id": 0, "booking_status": 1, "payment_status": 1, "total_amount": 1, "check_in_date": 1, "check_out_date": 1, "created_at": 1}
        ).to_list(length=10000)

        def _date_value(value):
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            if isinstance(value, str):
                try:
                    return date.fromisoformat(value[:10])
                except ValueError:
                    return None
            return None

        bookings_today = len([booking for booking in bookings if _date_value(booking.get("created_at")) == today])
        bookings_this_month = len([
            booking for booking in bookings
            if (created := _date_value(booking.get("created_at"))) and month_start <= created < next_month
        ])
        upcoming_checkins = len([
            booking for booking in bookings
            if booking.get("booking_status") in {"confirmed", "soft_lock"} and _date_value(booking.get("check_in_date")) and today <= _date_value(booking.get("check_in_date")) <= today + timedelta(days=7)
        ])
        upcoming_checkouts = len([
            booking for booking in bookings
            if booking.get("booking_status") in {"confirmed", "completed"} and _date_value(booking.get("check_out_date")) and today <= _date_value(booking.get("check_out_date")) <= today + timedelta(days=7)
        ])
        paid_bookings = [
            booking for booking in bookings
            if booking.get("payment_status") in {"paid", "partially_paid"} or booking.get("booking_status") in {"confirmed", "completed"}
        ]
        revenue_generated = round(sum(float(booking.get("total_amount") or 0) for booking in paid_bookings), 2)

        commissions = await db.commissions.find(
            {"broker_id": {"$in": my_broker_ids}},
            {"_id": 0, "commission_amount": 1}
        ).to_list(length=5000)
        broker_commission_generated = round(sum(float(item.get("commission_amount") or 0) for item in commissions), 2)

        rated_properties = [float(prop.get("rating") or 0) for prop in properties if float(prop.get("rating") or 0) > 0]
        average_property_rating = round(sum(rated_properties) / len(rated_properties), 1) if rated_properties else 0
        average_occupancy = round((len([b for b in bookings if b.get("booking_status") in {"confirmed", "completed"}]) / max(live_properties * 30, 1)) * 100, 1)

        stale_cutoff = now - timedelta(hours=48)
        pending_escalations = await db.property_verifications.count_documents({
            "broker_id": {"$in": my_broker_ids},
            "status": VerificationStatus.COMPLETED.value,
            "rm_reviewed": False,
            "completed_at": {"$lt": stale_cutoff}
        })
        sla_breaches = pending_escalations

        return {
            "scope": {
                "type": "branch_manager" if is_branch_manager else "rm",
                "label": "Branch Manager" if is_branch_manager else "RM",
            },
            "rms": {
                "total": len(branch_manager_rms),
                "active": len([rm for rm in branch_manager_rms if rm.get("is_active", True)]),
                "inactive": len([rm for rm in branch_manager_rms if not rm.get("is_active", True)]),
            },
            "brokers": {
                "total": total_brokers,
                "active": active_brokers,
                "inactive": inactive_brokers
            },
            "hosts": {
                "total": len(my_host_ids),
                "pending_verification": pending_host_verification
            },
            "properties": {
                "total": len(property_ids),
                "live": live_properties,
                "pending_verification": pending_property_verification,
                "rejected": rejected_properties,
                "draft": draft_properties,
                "average_rating": average_property_rating
            },
            "bookings": {
                "today": bookings_today,
                "this_month": bookings_this_month,
                "upcoming_checkins": upcoming_checkins,
                "upcoming_checkouts": upcoming_checkouts,
                "total": len(bookings)
            },
            "finance": {
                "revenue_generated": revenue_generated,
                "broker_commission_generated": broker_commission_generated
            },
            "performance": {
                "average_occupancy": average_occupancy,
                "average_property_rating": average_property_rating,
                "pending_escalations": pending_escalations,
                "sla_breaches": sla_breaches
            },
            "verifications": {
                "pending_review": pending_verifications,
                "under_review": properties_under_review
            },
            "subscriptions": {
                "expiring_soon": expiring_subscriptions
            }
        }
    
    except Exception as e:
        logger.error(f"Error fetching employee dashboard stats: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch dashboard stats"
        )


# ========== VERIFICATION REVIEW ==========

@router.get("/verifications/pending")
async def get_pending_verifications(
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Get all verifications pending RM/BM review."""
    try:
        broker_ids, _, property_query = await _get_rm_property_query(db, current_user["user_id"])
        rm_identifiers = await _get_rm_identifiers(db, current_user)
        
        profile = await _get_employee_profile(db, current_user)
        is_bm = _is_branch_manager_profile(profile)
        properties = await db.properties.find(property_query, {"_id": 0, "property_id": 1}).to_list(length=5000)
        property_ids = [prop["property_id"] for prop in properties if prop.get("property_id")]
        cursor = db.property_verifications.find(
            _pending_review_query(property_ids, rm_identifiers, is_bm),
            {"_id": 0}
        ).sort("completed_at", -1)
        
        raw_verifications = await cursor.to_list(length=300)
        
        # Enrich with property and broker details
        verifications = []
        for verification in raw_verifications:
            # Get property details
            property_data = await db.properties.find_one(
                {"property_id": verification["property_id"]},
                {"_id": 0, "title": 1, "address": 1, "city": 1, "images": 1, "bhk_type": 1, "owner_id": 1, "broker_id": 1, "rm_id": 1, "branch_manager_id": 1, "branch_manager_code": 1}
            )
            if property_data:
                verification["property_details"] = property_data
            owner_data, assignment = await _property_owner_assignment(db, property_data or {})
            assigned_bm = assignment.get("branch_manager_id") or assignment.get("branch_manager_code")
            assigned_rm = assignment.get("rm_id")
            assigned_broker = assignment.get("broker_id")
            if is_bm:
                bm_matches = assigned_bm in rm_identifiers or verification.get("branch_manager_id") in rm_identifiers
                rm_step_done = (
                    (verification.get("rm_reviewed") and verification.get("rm_approved") is True)
                    or (not assigned_broker and assigned_rm and verification.get("status") == VerificationStatus.COMPLETED.value)
                )
                if not bm_matches or not rm_step_done:
                    continue
            else:
                if assigned_rm not in rm_identifiers or assigned_broker == "":
                    continue
            
            # Get broker/RM who did the visit
            broker_id_to_fetch = assigned_broker or ""
            if is_bm and not assigned_broker:
                broker_id_to_fetch = assigned_rm or verification.get("rm_id") or ""

            broker_data = None
            if broker_id_to_fetch:
                broker_data = await db.users.find_one(
                    {"user_id": broker_id_to_fetch},
                    {"_id": 0, "full_name": 1, "lg_code": 1, "phone": 1, "employee_code": 1}
                )
            if broker_data:
                # Fallback code display if lg_code is missing
                if not broker_data.get("lg_code"):
                    broker_data["lg_code"] = broker_data.get("employee_code") or "N/A"
                verification["broker_details"] = broker_data
            verification["broker_id"] = assigned_broker or ""
            verification["rm_id"] = assigned_rm or verification.get("rm_id") or ""
            verification["branch_manager_id"] = assigned_bm or verification.get("branch_manager_id") or ""
            verification["review_stage"] = "branch_manager" if is_bm else "rm"
            verifications.append(verification)
            if len(verifications) >= 100:
                break
        
        return {
            "verifications": verifications,
            "total": len(verifications)
        }
    
    except Exception as e:
        logger.error(f"Error fetching pending verifications: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch pending verifications"
        )

@router.get("/verifications/history")
async def get_verification_history(
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Get all verifications reviewed by this RM/BM or rejected by admin."""
    try:
        broker_ids, _, property_query = await _get_rm_property_query(db, current_user["user_id"])
        rm_identifiers = await _get_rm_identifiers(db, current_user)
        
        profile = await _get_employee_profile(db, current_user)
        is_bm = _is_branch_manager_profile(profile)
        properties = await db.properties.find(property_query, {"_id": 0, "property_id": 1}).to_list(length=5000)
        property_ids = [prop["property_id"] for prop in properties if prop.get("property_id")]
        
        cursor = db.property_verifications.find(
            _history_review_query(property_ids, rm_identifiers, is_bm),
            {"_id": 0}
        ).sort("updated_at", -1)
        
        verifications = await cursor.to_list(length=100)
        
        # Enrich with property and broker details
        for verification in verifications:
            property_data = await db.properties.find_one(
                {"property_id": verification["property_id"]},
                {"_id": 0}
            )
            if property_data:
                verification["property_details"] = property_data
            
            broker_id_to_fetch = verification.get("broker_id")
            if not broker_id_to_fetch:
                broker_id_to_fetch = property_data.get("rm_id") if property_data else None

            broker_data = None
            if broker_id_to_fetch:
                broker_data = await db.users.find_one(
                    {"user_id": broker_id_to_fetch},
                    {"_id": 0, "full_name": 1, "lg_code": 1, "phone": 1, "employee_code": 1}
                )
            if broker_data:
                if not broker_data.get("lg_code"):
                    broker_data["lg_code"] = broker_data.get("employee_code") or "N/A"
                verification["broker_details"] = broker_data
        
        return {
            "verifications": verifications,
            "total": len(verifications)
        }
    except Exception as e:
        logger.error(f"Error fetching verification history: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch verification history"
        )


@router.get("/verifications/{verification_id}")
async def get_verification_details(
    verification_id: str,
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Get detailed verification report for review."""
    try:
        verification = await db.property_verifications.find_one(
            {"verification_id": verification_id},
            {"_id": 0}
        )
        
        if not verification:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Verification not found"
            )
        
        broker_ids, _, property_query = await _get_rm_property_query(db, current_user["user_id"])
        rm_identifiers = await _get_rm_identifiers(db, current_user)
        profile = await _get_employee_profile(db, current_user)
        is_bm = _is_branch_manager_profile(profile)
        properties = await db.properties.find(property_query, {"_id": 0, "property_id": 1}).to_list(length=5000)
        property_ids = [prop["property_id"] for prop in properties if prop.get("property_id")]
        branch_manager_match = any(
            verification.get("branch_manager_id") == identifier
            for identifier in rm_identifiers
        )
        if verification.get("broker_id") not in broker_ids and verification.get("property_id") not in property_ids and verification.get("rm_id") not in rm_identifiers and not branch_manager_match:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You are not authorized to view this verification report"
            )
        
        # Get full property details
        property_data = await db.properties.find_one(
            {"property_id": verification["property_id"]},
            {"_id": 0}
        )
        verification["property_details"] = property_data
        owner_data, assignment = await _property_owner_assignment(db, property_data or {})
        assigned_broker = assignment.get("broker_id")
        assigned_rm = assignment.get("rm_id")
        assigned_bm = assignment.get("branch_manager_id") or assignment.get("branch_manager_code")
        verification["broker_id"] = assigned_broker or ""
        verification["rm_id"] = assigned_rm or verification.get("rm_id") or ""
        verification["branch_manager_id"] = assigned_bm or verification.get("branch_manager_id") or ""
        
        # Get broker details
        broker_lookup_id = assigned_broker or (assigned_rm if is_bm else verification.get("broker_id"))
        broker_data = await db.users.find_one(
            {"user_id": broker_lookup_id},
            {"_id": 0, "full_name": 1, "lg_code": 1, "phone": 1, "email": 1, "employee_code": 1}
        ) if broker_lookup_id else None
        if broker_data and not broker_data.get("lg_code"):
            broker_data["lg_code"] = broker_data.get("employee_code") or "N/A"
        verification["broker_details"] = broker_data
        
        # Get owner details
        owner_data = owner_data or await db.users.find_one(
            {"user_id": verification["owner_id"]},
            {"_id": 0, "full_name": 1, "phone": 1, "email": 1}
        )
        verification["owner_details"] = owner_data
        verification["review_stage"] = "branch_manager" if is_bm else "rm"
        
        return verification
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching verification details: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch verification details"
        )

@router.get("/verifications/{verification_id}/export-report")
async def export_verification_report_xlsx(
    verification_id: str,
    current_user: dict = Depends(get_current_user),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Export a single verification report as a Luxury Premium XLSX."""
    if current_user["role"] not in (UserRole.EMPLOYEE.value, "admin"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Employee or Admin access required"
        )
    print(f"DEBUG: Generating Luxury Report for {verification_id}")
    try:
        # Get full verification details
        verification = await db.property_verifications.find_one(
            {"verification_id": verification_id},
            {"_id": 0}
        )
        
        if not verification:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Verification not found"
            )

        if current_user["role"] == UserRole.EMPLOYEE.value:
            broker_ids, _ = await _get_rm_scope(db, current_user["user_id"])
            if verification.get("broker_id") not in broker_ids:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="You are not authorized to export this verification report"
                )
        
        # Enrich with property, broker and owner
        property_data = await db.properties.find_one({"property_id": verification["property_id"]}, {"_id": 0})
        broker_data = await db.users.find_one({"user_id": verification["broker_id"]}, {"_id": 0, "full_name": 1, "lg_code": 1})
        owner_data = await db.users.find_one({"user_id": verification["owner_id"]}, {"_id": 0, "full_name": 1})

        # Create Styled Excel Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Executive Audit"
        ws.sheet_view.showGridLines = False

        # Luxury Noir Palette (Black & Gold)
        noir = "1A1A1A"
        gold = "D4AF37"
        platinum = "E5E4E2"
        white = "FFFFFF"
        
        # Styles
        title_font = Font(name='Georgia', size=24, bold=True, color=gold)
        header_font = Font(name='Georgia', size=14, bold=True, color=gold)
        label_font = Font(name='Arial', size=11, bold=True, color=platinum)
        value_font = Font(name='Arial', size=11, color=white)
        
        bg_fill = PatternFill(start_color=noir, end_color=noir, fill_type="solid")
        
        gold_side = Side(style='thin', color=gold)
        gold_border = Border(left=gold_side, right=gold_side, top=gold_side, bottom=gold_side)
        
        # Fill background
        for r in range(1, 100):
            for c in range(1, 10):
                ws.cell(row=r, column=c).fill = bg_fill

        # Set Column Widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 30

        # 1. Luxury Header
        ws.merge_cells('A1:C3')
        cell = ws['A1']
        cell.value = "X-SPACE360 | ELITE AUDIT"
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=gold_side, right=gold_side, top=gold_side, bottom=gold_side)

        curr_row = 5
        
        # 2. Property Hero
        ws.merge_cells(f'A{curr_row}:C{curr_row+1}')
        hero = ws[f'A{curr_row}']
        hero.value = property_data.get("title", "N/A").upper()
        hero.font = Font(name='Arial', size=16, bold=True, color=white)
        hero.alignment = Alignment(horizontal='center', vertical='center')
        hero.border = gold_border
        curr_row += 4

        # 3. Status Badge
        ws[f'A{curr_row}'] = "AUDIT STATUS"
        ws[f'B{curr_row}'] = "CERTIFIED" if verification.get("rm_approved") else "UNDER REVIEW"
        ws[f'A{curr_row}'].font = label_font
        ws[f'B{curr_row}'].font = Font(name='Arial', size=12, bold=True, color=noir)
        ws[f'B{curr_row}'].fill = PatternFill(start_color=gold, end_color=gold, fill_type="solid")
        ws[f'B{curr_row}'].alignment = Alignment(horizontal='center')
        ws[f'A{curr_row}'].border = gold_border
        ws[f'B{curr_row}'].border = gold_border
        curr_row += 3

        # 4. Details Section
        ws.merge_cells(f'A{curr_row}:C{curr_row}')
        ws[f'A{curr_row}'] = "I. METADATA & AUTHENTICATION"
        ws[f'A{curr_row}'].font = header_font
        ws[f'A{curr_row}'].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        curr_row += 1

        details = [
            ("Verification ID", verification["verification_id"]),
            ("Audit Reference", datetime.now(timezone.utc).strftime('%d %B %Y')),
            ("Lead Auditor", current_user.get("full_name", "N/A")),
            ("Field Intelligence", broker_data.get("full_name", "N/A"))
        ]

        for label, val in details:
            ws[f'A{curr_row}'] = label
            ws[f'B{curr_row}'] = val
            ws[f'A{curr_row}'].font = label_font
            ws[f'B{curr_row}'].font = value_font
            ws[f'A{curr_row}'].border = gold_border
            ws[f'B{curr_row}'].border = gold_border
            curr_row += 1
        
        curr_row += 2

        # 5. Checklist Section
        ws.merge_cells(f'A{curr_row}:C{curr_row}')
        cell = ws[f'A{curr_row}']
        cell.value = "II. COMPLIANCE CHECKLIST"
        cell.font = header_font
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        curr_row += 1

        checklist = verification.get("checklist", {})
        for key, value in checklist.items():
            ws[f'A{curr_row}'] = key.replace("_", " ").title()
            ws[f'B{curr_row}'] = "✔ COMPLIANT" if value else "✘ DEFICIENT"
            ws[f'A{curr_row}'].font = value_font
            ws[f'B{curr_row}'].font = Font(name='Arial', size=11, bold=True, color="27AE60" if value else "C0392B")
            ws[f'A{curr_row}'].border = gold_border
            ws[f'B{curr_row}'].border = gold_border
            curr_row += 1

        # Footer
        curr_row += 3
        ws[f'A{curr_row}'] = "OFFICIAL ELECTRONIC RECORD"
        ws[f'A{curr_row}'].font = Font(size=8, color=gold)
        ws[f'C{curr_row}'] = "SECURE DOCUMENT"
        ws[f'C{curr_row}'].font = Font(size=8, italic=True, color=gold)
        ws[f'C{curr_row}'].alignment = Alignment(horizontal='right')

        # Save to IO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Use a timestamp to prevent browser caching
        ts = datetime.now().strftime('%H%M%S')
        filename = f"Elite_Report_{verification_id}_{ts}.xlsx"
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

        # Save to IO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Verification_Report_{verification_id}.xlsx"
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        logger.error(f"Error exporting verification XLSX: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export report: {str(e)}"
        )

@router.post("/verifications/{verification_id}/approve")
async def approve_verification(
    verification_id: str,
    payload: Optional[RMReviewRequest] = None,
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Approve verification and forward to admin."""
    try:
        remarks = payload.remarks if payload else None
        verification = await db.property_verifications.find_one(
            {"verification_id": verification_id}
        )
        
        if not verification:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Verification not found"
            )
            
        rm_identifiers = await _get_rm_identifiers(db, current_user)
        profile = await _get_employee_profile(db, current_user)
        is_bm = _is_branch_manager_profile(profile)
        _, _, property_query = await _get_rm_property_query(db, current_user["user_id"])
        properties = await db.properties.find(property_query, {"_id": 0, "property_id": 1}).to_list(length=5000)
        property_ids = [prop["property_id"] for prop in properties if prop.get("property_id")]
        property_data = await db.properties.find_one({"property_id": verification.get("property_id")}, {"_id": 0}) or {}
        _, assignment = await _property_owner_assignment(db, property_data)
        assigned_bm = assignment.get("branch_manager_id") or assignment.get("branch_manager_code") or verification.get("branch_manager_id")
        assigned_rm = assignment.get("rm_id") or verification.get("rm_id")
        assigned_to_current_stage = (
            any(assigned_bm == identifier for identifier in rm_identifiers)
            if is_bm
            else assigned_rm in rm_identifiers
        )
        if verification.get("property_id") not in property_ids and not assigned_to_current_stage:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You are not authorized to approve this verification"
            )
        
        now = datetime.now(timezone.utc)
        if is_bm:
            rm_step_completed = (
                (verification.get("rm_reviewed") and verification.get("rm_approved") is True)
                or not assignment.get("broker_id")
            )
            if not rm_step_completed:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="RM verification must be completed before Branch Manager approval"
                )
            await db.property_verifications.update_one(
                {"verification_id": verification_id},
                {"$set": {
                    "rm_reviewed": True,
                    "rm_approved": True,
                    "rm_id": assigned_rm or verification.get("rm_id") or "",
                    "branch_manager_reviewed": True,
                    "branch_manager_approved": True,
                    "branch_manager_remarks": remarks,
                    "branch_manager_id": current_user["user_id"],
                    "branch_manager_reviewed_at": now,
                    "reviewed_at": now,
                    "updated_at": now
                }}
            )
        else:
            await db.property_verifications.update_one(
                {"verification_id": verification_id},
                {"$set": {
                    "rm_reviewed": True,
                    "rm_approved": True,
                    "rm_remarks": remarks,
                    "rm_id": current_user["user_id"],
                    "reviewed_at": now,
                    "updated_at": now
                }}
            )
        
        # Property stays in under_review - awaiting admin final approval
        await db.properties.update_one(
            {"property_id": verification["property_id"]},
            {"$set": {
                "status": "under_review",
                "updated_at": datetime.now(timezone.utc)
            }}
        )

        try:
            await write_audit_log(
                db,
                user_id=current_user["user_id"],
                role=current_user.get("role"),
                module="rm_verification",
                action="branch_manager_verification_approved" if is_bm else "rm_verification_approved",
                record_id=verification_id,
                old_value={
                    "rm_reviewed": verification.get("rm_reviewed"),
                    "rm_approved": verification.get("rm_approved"),
                    "status": verification.get("status"),
                    "property_id": verification.get("property_id"),
                },
                new_value={
                    "rm_reviewed": True if not is_bm else verification.get("rm_reviewed"),
                    "rm_approved": True if not is_bm else verification.get("rm_approved"),
                    "branch_manager_reviewed": True if is_bm else verification.get("branch_manager_reviewed"),
                    "branch_manager_approved": True if is_bm else verification.get("branch_manager_approved"),
                    "status": VerificationStatus.COMPLETED.value,
                    "property_status": "under_review",
                    "remarks": remarks,
                },
                reason=remarks or "",
            )
        except Exception as audit_err:
            logger.warning(f"Failed to write RM approval audit: {audit_err}")

        # Notify admins + host
        should_notify_admin = is_bm or not assigned_bm
        if should_notify_admin:
            try:
                from services.verification_workflow import on_rm_decision
                verification_doc = await db.property_verifications.find_one(
                    {"verification_id": verification_id}, {"_id": 0}
                )
                asyncio.create_task(on_rm_decision(db, verification_doc, approved=True, remarks=remarks or ""))
            except Exception as wf_err:
                logger.warning(f"on_rm_decision (approve) trigger failed: {wf_err}")

        logger.info(f"Verification {verification_id} approved by {'BM' if is_bm else 'RM'} {current_user['user_id']}")
        return {
            "message": "Verification approved. Forwarded to admin for final approval." if should_notify_admin else "Verification approved. Forwarded to Branch Manager for review.",
            "verification_id": verification_id
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error approving verification: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to approve verification"
        )

@router.post("/verifications/{verification_id}/reject")
async def reject_verification(
    verification_id: str,
    payload: RMRejectRequest,
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Reject verification and send back to host."""
    try:
        reason = payload.reason
        verification = await db.property_verifications.find_one(
            {"verification_id": verification_id}
        )
        
        if not verification:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Verification not found"
            )
            
        rm_identifiers = await _get_rm_identifiers(db, current_user)
        profile = await _get_employee_profile(db, current_user)
        is_bm = _is_branch_manager_profile(profile)
        _, _, property_query = await _get_rm_property_query(db, current_user["user_id"])
        properties = await db.properties.find(property_query, {"_id": 0, "property_id": 1}).to_list(length=5000)
        property_ids = [prop["property_id"] for prop in properties if prop.get("property_id")]
        property_data = await db.properties.find_one({"property_id": verification.get("property_id")}, {"_id": 0}) or {}
        _, assignment = await _property_owner_assignment(db, property_data)
        assigned_bm = assignment.get("branch_manager_id") or assignment.get("branch_manager_code") or verification.get("branch_manager_id")
        assigned_rm = assignment.get("rm_id") or verification.get("rm_id")
        assigned_to_current_stage = (
            any(assigned_bm == identifier for identifier in rm_identifiers)
            if is_bm
            else assigned_rm in rm_identifiers
        )
        if verification.get("property_id") not in property_ids and not assigned_to_current_stage:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You are not authorized to reject this verification"
            )
        
        now = datetime.now(timezone.utc)
        if is_bm:
            update_fields = {
                "branch_manager_reviewed": True,
                "branch_manager_approved": False,
                "branch_manager_remarks": reason,
                "branch_manager_id": current_user["user_id"],
                "branch_manager_reviewed_at": now,
                "status": VerificationStatus.REJECTED.value,
                "reviewed_at": now,
                "updated_at": now
            }
        else:
            update_fields = {
                "rm_reviewed": True,
                "rm_approved": False,
                "rm_remarks": reason,
                "rm_id": current_user["user_id"],
                "status": VerificationStatus.REJECTED.value,
                "reviewed_at": now,
                "updated_at": now
            }

        await db.property_verifications.update_one(
            {"verification_id": verification_id},
            {"$set": update_fields}
        )
        
        # Update property status back to needs resubmission
        await db.properties.update_one(
            {"property_id": verification["property_id"]},
            {"$set": {
                "status": "draft",
                "verification_remarks": reason,
                "updated_at": now
            }}
        )

        try:
            await write_audit_log(
                db,
                user_id=current_user["user_id"],
                role=current_user.get("role"),
                module="rm_verification",
                action="branch_manager_verification_rejected" if is_bm else "rm_verification_rejected",
                record_id=verification_id,
                old_value={
                    "rm_reviewed": verification.get("rm_reviewed"),
                    "rm_approved": verification.get("rm_approved"),
                    "branch_manager_reviewed": verification.get("branch_manager_reviewed"),
                    "branch_manager_approved": verification.get("branch_manager_approved"),
                    "status": verification.get("status"),
                    "property_id": verification.get("property_id"),
                },
                new_value={
                    **update_fields,
                    "property_status": "draft",
                    "reason": reason,
                },
                reason=reason,
            )
        except Exception as audit_err:
            logger.warning(f"Failed to write verification rejection audit: {audit_err}")

        # Notify host that the listing needs revision
        try:
            from services.verification_workflow import on_rm_decision
            verification_doc = await db.property_verifications.find_one(
                {"verification_id": verification_id}, {"_id": 0}
            )
            asyncio.create_task(on_rm_decision(db, verification_doc, approved=False, remarks=reason))
        except Exception as wf_err:
            logger.warning(f"on_rm_decision (reject) trigger failed: {wf_err}")

        logger.info(f"Verification {verification_id} rejected by {'BM' if is_bm else 'RM'} {current_user['user_id']}")
        return {
            "message": "Verification rejected. Host will be notified to resubmit.",
            "verification_id": verification_id
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error rejecting verification: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to reject verification"
        )

# ========== BROKER OVERSIGHT ==========

@router.get("/brokers")
async def get_all_brokers(
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Get all brokers under this RM."""
    try:
        # Get brokers assigned to this RM
        broker_ids, _ = await _get_rm_scope(db, current_user)
        cursor = db.users.find(
            {"role": "broker", "user_id": {"$in": broker_ids}},
            {"_id": 0, "password_hash": 0}
        )
        brokers = await cursor.to_list(length=200)
        
        # Get stats for each broker
        for broker in brokers:
            broker_id = broker["user_id"]
            
            # Count owners
            owner_count = await db.users.count_documents({
                "broker_id": broker_id,
                "role": "host"
            })
            
            # Count properties
            property_count = await db.properties.count_documents({"broker_id": broker_id})
            live_properties = await db.properties.count_documents({
                "broker_id": broker_id,
                "status": "live"
            })
            broker_property_ids = [
                item["property_id"]
                for item in await db.properties.find({"broker_id": broker_id}, {"_id": 0, "property_id": 1}).to_list(length=1000)
                if item.get("property_id")
            ]
            
            # Pending verifications
            pending_verifications = await db.property_verifications.count_documents({
                "broker_id": broker_id,
                "$or": [
                    {"status": {"$in": ["pending", "in_progress", VerificationStatus.COMPLETED.value]}},
                    {"rm_reviewed": False}
                ]
            })
            bookings = await db.bookings.find(
                {
                    "$or": [
                        {"broker_id": broker_id},
                        {"property_id": {"$in": broker_property_ids}},
                    ]
                },
                {"_id": 0, "booking_status": 1, "payment_status": 1, "total_amount": 1, "created_at": 1}
            ).to_list(length=2000)
            revenue_generated = round(sum(
                float(booking.get("total_amount") or 0)
                for booking in bookings
                if booking.get("payment_status") in {"paid", "partially_paid"} or booking.get("booking_status") in {"confirmed", "completed"}
            ), 2)
            commissions = await db.commissions.find(
                {"broker_id": broker_id},
                {"_id": 0, "commission_amount": 1, "payment_status": 1}
            ).to_list(length=1000)
            commission_earned = round(sum(float(item.get("commission_amount") or 0) for item in commissions), 2)
            stale_cutoff = datetime.now(timezone.utc) - timedelta(hours=48)
            pending_escalations = await db.property_verifications.count_documents({
                "broker_id": broker_id,
                "rm_reviewed": False,
                "created_at": {"$lt": stale_cutoff}
            })
            latest_audit = await db.audit_logs.find_one(
                {
                    "$or": [
                        {"user_id": broker_id},
                        {"record_id": broker_id},
                        {"record_id": {"$in": broker_property_ids}},
                    ]
                },
                {"_id": 0, "created_at": 1, "action": 1, "module": 1},
                sort=[("created_at", -1)]
            )
            performance_rating = round(
                min(5, (live_properties * 0.4) + (len(bookings) * 0.15) + (revenue_generated / 100000)),
                1
            ) if property_count or bookings else 0
            
            broker["stats"] = {
                "owners": owner_count,
                "hosts": owner_count,
                "properties": property_count,
                "live_properties": live_properties,
                "bookings": len(bookings),
                "revenue_generated": revenue_generated,
                "commission_earned": commission_earned,
                "pending_verifications": pending_verifications,
                "pending_escalations": pending_escalations,
                "performance_rating": performance_rating,
                "last_activity": latest_audit
            }
        
        return {
            "brokers": brokers,
            "total": len(brokers)
        }
    
    except Exception as e:
        logger.error(f"Error fetching brokers: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch brokers"
        )

@router.get("/brokers/{broker_id}/portfolio")
async def get_broker_portfolio(
    broker_id: str,
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Get detailed portfolio for a specific broker."""
    try:
        # Get broker details
        broker = await db.users.find_one(
            {"user_id": broker_id, "role": "broker"},
            {"_id": 0, "password_hash": 0}
        )
        
        if not broker:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Broker not found"
            )
            
        broker_ids, _ = await _get_rm_scope(db, current_user)
        if broker_id not in broker_ids:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="This broker is not assigned to you"
            )
        
        # Get all properties
        property_cursor = db.properties.find({"broker_id": broker_id}, {"_id": 0})
        properties = await property_cursor.to_list(length=200)
        property_ids = [prop["property_id"] for prop in properties if prop.get("property_id")]
        
        # Get all owners
        owner_cursor = db.users.find(
            {"broker_id": broker_id, "role": "host"},
            {"_id": 0, "password_hash": 0}
        )
        owners = await owner_cursor.to_list(length=200)

        bookings = await db.bookings.find(
            {
                "$or": [
                    {"broker_id": broker_id},
                    {"property_id": {"$in": property_ids}},
                ]
            },
            {"_id": 0}
        ).sort("created_at", -1).to_list(length=200)
        commissions = await db.commissions.find({"broker_id": broker_id}, {"_id": 0}).sort("created_at", -1).to_list(length=200)
        verifications = await db.property_verifications.find({"broker_id": broker_id}, {"_id": 0}).sort("updated_at", -1).to_list(length=100)
        audit_logs = await db.audit_logs.find(
            {
                "$or": [
                    {"user_id": broker_id},
                    {"record_id": broker_id},
                    {"record_id": {"$in": property_ids}},
                ]
            },
            {"_id": 0}
        ).sort("created_at", -1).limit(50).to_list(length=50)

        paid_bookings = [
            booking for booking in bookings
            if booking.get("payment_status") in {"paid", "partially_paid"} or booking.get("booking_status") in {"confirmed", "completed"}
        ]
        stale_cutoff = datetime.now(timezone.utc) - timedelta(hours=48)
        pending_escalations = [
            item for item in verifications
            if not item.get("rm_reviewed") and item.get("created_at") and item.get("created_at") < stale_cutoff
        ]
        summary = {
            "hosts": len(owners),
            "properties": len(properties),
            "live_properties": len([prop for prop in properties if prop.get("status") == "live"]),
            "pending_properties": len([prop for prop in properties if prop.get("status") in {"draft", "pending_verification", "under_review"}]),
            "bookings": len(bookings),
            "revenue_generated": round(sum(float(booking.get("total_amount") or 0) for booking in paid_bookings), 2),
            "commission_earned": round(sum(float(item.get("commission_amount") or 0) for item in commissions), 2),
            "pending_verifications": len([item for item in verifications if not item.get("rm_reviewed")]),
            "pending_escalations": len(pending_escalations),
        }
        
        return {
            "broker": broker,
            "properties": properties,
            "owners": owners,
            "hosts": owners,
            "bookings": bookings,
            "commissions": commissions,
            "verifications": verifications,
            "audit_logs": audit_logs,
            "summary": summary
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching broker portfolio: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch broker portfolio"
        )

# ========== HOST MANAGEMENT ==========

@router.get("/hosts")
async def get_rm_hosts(
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Get all hosts under this RM's assigned brokers."""
    try:
        broker_ids, scoped_host_ids = await _get_rm_scope(db, current_user["user_id"])
        host_scope_or = []
        if broker_ids:
            host_scope_or.append({"broker_id": {"$in": broker_ids}})
        if scoped_host_ids:
            host_scope_or.append({"user_id": {"$in": scoped_host_ids}})
        if not host_scope_or:
            return {"hosts": [], "total": 0}
        cursor = db.users.find(
            {"role": "host", "$or": host_scope_or},
            {"_id": 0, "password_hash": 0}
        ).sort("created_at", -1)
        hosts = await cursor.to_list(length=500)
        broker_map = {
            broker["user_id"]: broker
            for broker in await db.users.find(
                {"user_id": {"$in": broker_ids}},
                {"_id": 0, "user_id": 1, "full_name": 1, "lg_code": 1}
            ).to_list(length=1000)
        }

        for host in hosts:
            host_id = host["user_id"]
            broker = broker_map.get(host.get("broker_id"), {})
            properties = await db.properties.find(
                {"owner_id": host_id},
                {"_id": 0, "property_id": 1, "status": 1}
            ).to_list(length=500)
            bookings = await db.bookings.find(
                {"host_id": host_id},
                {"_id": 0, "booking_status": 1, "payment_status": 1, "total_amount": 1}
            ).to_list(length=1000)
            paid_bookings = [
                booking for booking in bookings
                if booking.get("payment_status") in {"paid", "partially_paid"} or booking.get("booking_status") in {"confirmed", "completed"}
            ]
            host["broker_name"] = broker.get("full_name") or "Not assigned"
            host["broker_lg_code"] = broker.get("lg_code") or host.get("lg_code")
            host["total_properties"] = len(properties)
            host["live_properties"] = len([prop for prop in properties if prop.get("status") == "live"])
            host["pending_properties"] = len([prop for prop in properties if prop.get("status") in {"draft", "pending", "pending_verification", "under_review", "rejected"}])
            host["total_bookings"] = len(bookings)
            host["revenue_generated"] = round(sum(float(booking.get("total_amount") or 0) for booking in paid_bookings), 2)
            host["verification_status"] = host.get("kyc_status") or "pending"

        return {"hosts": hosts, "total": len(hosts)}

    except Exception as e:
        logger.error(f"Error fetching RM hosts: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch hosts"
        )


@router.get("/hosts/{host_id}/details")
async def get_rm_host_details(
    host_id: str,
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Get detailed host view for RM scoped host management."""
    try:
        broker_ids, host_ids = await _get_rm_scope(db, current_user["user_id"])
        if host_id not in host_ids:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="This host is not assigned to you"
            )

        host = await db.users.find_one(
            {"user_id": host_id, "role": "host"},
            {"_id": 0, "password_hash": 0}
        )
        if not host:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Host not found")

        broker = None
        if host.get("broker_id"):
            broker = await db.users.find_one(
                {"user_id": host.get("broker_id"), "role": "broker"},
                {"_id": 0, "password_hash": 0}
            )

        properties = await db.properties.find({"owner_id": host_id}, {"_id": 0}).sort("created_at", -1).to_list(length=200)
        bookings = await db.bookings.find({"host_id": host_id}, {"_id": 0}).sort("created_at", -1).to_list(length=100)
        payments = await db.transactions.find(
            {"$or": [{"host_id": host_id}, {"user_id": host_id}]},
            {"_id": 0}
        ).sort("created_at", -1).to_list(length=100)
        verifications = await db.property_verifications.find({"owner_id": host_id}, {"_id": 0}).sort("created_at", -1).to_list(length=100)
        audit_events = await db.audit_logs.find(
            {"$or": [{"user_id": host_id}, {"entity_id": host_id}, {"target_id": host_id}, {"record_id": host_id}]},
            {"_id": 0}
        ).sort("created_at", -1).to_list(length=50)

        return {
            "host": host,
            "owner": host,
            "broker": broker,
            "properties": properties,
            "bookings": bookings,
            "payments": payments,
            "verifications": verifications,
            "audit_events": audit_events,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching RM host details: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch host details"
        )

# ========== PROPERTY MANAGEMENT ==========

async def _get_rm_property_query(db: AsyncIOMotorDatabase, rm_id: str):
    broker_ids, host_ids = await _get_rm_scope(db, rm_id)
    identifiers = await _get_rm_identifiers(db, rm_id)
    property_or = _field_matches_identifiers("rm_id", identifiers)
    property_or.extend(_field_matches_identifiers("branch_manager_id", identifiers))
    if broker_ids:
        property_or.append({"broker_id": {"$in": broker_ids}})
    if host_ids:
        property_or.append({"owner_id": {"$in": host_ids}})
    return broker_ids, host_ids, {"$or": property_or}


def _property_verification_stage(property_doc: dict, verification: Optional[dict]) -> dict:
    status = property_doc.get("status") or "draft"
    verification = verification or {}
    rm_reviewed = bool(verification.get("rm_reviewed"))
    admin_reviewed = bool(verification.get("admin_reviewed"))
    stages = [
        ("Basic Information", "completed" if property_doc.get("title") else "pending"),
        ("Location", "completed" if property_doc.get("city") and property_doc.get("address") else "pending"),
        ("Amenities", "completed" if property_doc.get("amenities") else "pending"),
        ("Pricing", "completed" if property_doc.get("price_per_night") or property_doc.get("price_per_week") or property_doc.get("price_per_month") else "pending"),
        ("Images", "completed" if property_doc.get("images") else "pending"),
        ("Videos", "completed" if property_doc.get("video_url") or property_doc.get("youtube_short_url") or property_doc.get("youtube_long_url") else "waiting"),
        ("Documents", "completed" if verification.get("documents") or verification.get("checklist") else "pending" if status != "draft" else "waiting"),
        ("Broker Verification", "completed" if verification.get("status") == VerificationStatus.COMPLETED.value or verification.get("completed_at") else "pending" if status in {"pending_verification", "under_review"} else "waiting"),
        ("RM Verification", "rejected" if rm_reviewed and verification.get("rm_approved") is False else "completed" if rm_reviewed and verification.get("rm_approved") is True else "pending" if verification.get("status") == VerificationStatus.COMPLETED.value else "waiting"),
        ("Finance Approval", "pending" if rm_reviewed and verification.get("rm_approved") is True else "waiting"),
        ("Admin Approval", "rejected" if admin_reviewed and verification.get("admin_approved") is False else "completed" if admin_reviewed and verification.get("admin_approved") is True else "pending" if rm_reviewed and verification.get("rm_approved") is True else "waiting"),
        ("Property Live", "completed" if status == "live" else "rejected" if status == "rejected" else "pending" if admin_reviewed and verification.get("admin_approved") is True else "waiting"),
    ]
    completed = len([stage for stage in stages if stage[1] == "completed"])
    return {
        "current_stage": next((label for label, stage_status in stages if stage_status in {"pending", "rejected"}), "Property Live" if status == "live" else "Basic Information"),
        "completed": completed,
        "total": len(stages),
        "stages": [{"label": label, "status": stage_status} for label, stage_status in stages],
    }


@router.get("/properties")
async def get_rm_properties(
    status_filter: Optional[str] = None,
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Get all properties under this RM through broker/host/RM assignment."""
    try:
        _, _, property_query = await _get_rm_property_query(db, current_user["user_id"])
        if status_filter and status_filter != "all":
            property_query = {"$and": [property_query, {"status": status_filter}]}

        properties = await db.properties.find(property_query, {"_id": 0}).sort("updated_at", -1).to_list(length=500)
        broker_ids = list({prop.get("broker_id") for prop in properties if prop.get("broker_id")})
        host_ids = list({prop.get("owner_id") for prop in properties if prop.get("owner_id")})
        brokers = {
            row["user_id"]: row
            for row in await db.users.find({"user_id": {"$in": broker_ids}}, {"_id": 0, "user_id": 1, "full_name": 1, "lg_code": 1}).to_list(length=500)
        }
        hosts = {
            row["user_id"]: row
            for row in await db.users.find({"user_id": {"$in": host_ids}}, {"_id": 0, "user_id": 1, "full_name": 1, "kyc_status": 1}).to_list(length=500)
        }

        for prop in properties:
            prop_id = prop.get("property_id")
            verification = await db.property_verifications.find_one({"property_id": prop_id}, {"_id": 0}, sort=[("updated_at", -1)])
            booking_count = await db.bookings.count_documents({"property_id": prop_id})
            revenue_rows = await db.bookings.find(
                {"property_id": prop_id, "booking_status": {"$in": ["confirmed", "completed"]}},
                {"_id": 0, "total_amount": 1}
            ).to_list(length=1000)
            prop["host_summary"] = hosts.get(prop.get("owner_id"), {})
            prop["broker_summary"] = brokers.get(prop.get("broker_id"), {})
            prop["booking_count"] = booking_count
            prop["revenue_generated"] = round(sum(float(row.get("total_amount") or 0) for row in revenue_rows), 2)
            prop["verification_stage"] = _property_verification_stage(prop, verification)

        summary = {
            "total": len(properties),
            "live": len([prop for prop in properties if prop.get("status") == "live"]),
            "pending_verification": len([prop for prop in properties if prop.get("status") in {"pending_verification", "under_review"}]),
            "rejected": len([prop for prop in properties if prop.get("status") == "rejected"]),
            "draft": len([prop for prop in properties if prop.get("status") == "draft"]),
        }

        return {"properties": properties, "summary": summary, "total": len(properties)}

    except Exception as e:
        logger.error(f"Error fetching RM properties: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to fetch properties")


@router.get("/properties/{property_id}/details")
async def get_rm_property_details(
    property_id: str,
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Get detailed property view for RM scoped property operations."""
    try:
        _, _, property_query = await _get_rm_property_query(db, current_user["user_id"])
        scoped_query = {"$and": [property_query, {"property_id": property_id}]}
        prop = await db.properties.find_one(scoped_query, {"_id": 0})
        if not prop:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Property not found or not assigned to you")

        host = await db.users.find_one({"user_id": prop.get("owner_id")}, {"_id": 0, "password_hash": 0})
        broker = await db.users.find_one({"user_id": prop.get("broker_id")}, {"_id": 0, "password_hash": 0}) if prop.get("broker_id") else None
        verifications = await db.property_verifications.find({"property_id": property_id}, {"_id": 0}).sort("updated_at", -1).to_list(length=100)
        bookings = await db.bookings.find({"property_id": property_id}, {"_id": 0}).sort("created_at", -1).to_list(length=100)
        audit_logs = await db.audit_logs.find({"record_id": property_id}, {"_id": 0}).sort("created_at", -1).to_list(length=50)
        prop["verification_stage"] = _property_verification_stage(prop, verifications[0] if verifications else None)

        return {
            "property": prop,
            "host": host,
            "broker": broker,
            "verifications": verifications,
            "bookings": bookings,
            "audit_logs": audit_logs,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching RM property details: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to fetch property details")

# ========== BOOKING MANAGEMENT ==========

async def _get_rm_booking_query(db: AsyncIOMotorDatabase, rm_id: str):
    broker_ids, host_ids, property_query = await _get_rm_property_query(db, rm_id)
    identifiers = await _get_rm_identifiers(db, rm_id)
    properties = await db.properties.find(property_query, {"_id": 0, "property_id": 1}).to_list(length=5000)
    property_ids = [prop["property_id"] for prop in properties if prop.get("property_id")]
    return broker_ids, host_ids, property_ids, {
        "$or": [
            {"broker_id": {"$in": broker_ids}},
            {"user_broker_id": {"$in": broker_ids}},
            {"host_id": {"$in": host_ids}},
            {"owner_id": {"$in": host_ids}},
            {"user_id": {"$in": host_ids}},
            {"property_id": {"$in": property_ids}},
            *_field_matches_identifiers("rm_id", identifiers),
            *_field_matches_identifiers("user_rm_id", identifiers),
            *_field_matches_identifiers("employee_id", identifiers),
            *_field_matches_identifiers("assigned_rm_id", identifiers),
        ]
    }


def _booking_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value[:10])
        except ValueError:
            return None
    return None


@router.get("/bookings")
async def get_rm_bookings(
    status_filter: Optional[str] = None,
    period: Optional[str] = None,
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Get RM scoped bookings across assigned brokers, hosts and properties."""
    try:
        _, _, _, booking_query = await _get_rm_booking_query(db, current_user["user_id"])
        if status_filter and status_filter != "all":
            booking_query = {"$and": [booking_query, {"booking_status": status_filter}]}

        bookings = await db.bookings.find(booking_query, {"_id": 0}).sort("created_at", -1).to_list(length=500)
        today = date.today()
        if period and period != "all":
            if period == "daily":
                bookings = [booking for booking in bookings if _booking_date(booking.get("created_at")) == today]
            elif period == "weekly":
                start = today - timedelta(days=today.weekday())
                end = start + timedelta(days=7)
                bookings = [booking for booking in bookings if (created := _booking_date(booking.get("created_at"))) and start <= created < end]
            elif period == "monthly":
                start = today.replace(day=1)
                end = date(today.year + (1 if today.month == 12 else 0), 1 if today.month == 12 else today.month + 1, 1)
                bookings = [booking for booking in bookings if (created := _booking_date(booking.get("created_at"))) and start <= created < end]
            elif period == "yearly":
                start = date(today.year, 1, 1)
                end = date(today.year + 1, 1, 1)
                bookings = [booking for booking in bookings if (created := _booking_date(booking.get("created_at"))) and start <= created < end]

        property_ids = list({booking.get("property_id") for booking in bookings if booking.get("property_id")})
        host_ids = list({booking.get("host_id") for booking in bookings if booking.get("host_id")})
        broker_ids = list({booking.get("broker_id") for booking in bookings if booking.get("broker_id")})
        properties = {
            row["property_id"]: row
            for row in await db.properties.find({"property_id": {"$in": property_ids}}, {"_id": 0, "property_id": 1, "title": 1, "city": 1, "broker_id": 1, "owner_id": 1}).to_list(length=500)
        }
        hosts = {
            row["user_id"]: row
            for row in await db.users.find({"user_id": {"$in": host_ids}}, {"_id": 0, "user_id": 1, "full_name": 1, "phone": 1, "email": 1}).to_list(length=500)
        }
        brokers = {
            row["user_id"]: row
            for row in await db.users.find({"user_id": {"$in": broker_ids}}, {"_id": 0, "user_id": 1, "full_name": 1, "lg_code": 1}).to_list(length=500)
        }
        for booking in bookings:
            prop = properties.get(booking.get("property_id"), {})
            booking["property_summary"] = prop
            booking["host_summary"] = hosts.get(booking.get("host_id") or prop.get("owner_id"), {})
            booking["broker_summary"] = brokers.get(booking.get("broker_id") or prop.get("broker_id"), {})

        revenue_bookings = [
            booking for booking in bookings
            if booking.get("payment_status") in {"paid", "partially_paid"} or booking.get("booking_status") in {"confirmed", "completed"}
        ]
        summary = {
            "total": len(bookings),
            "pending": len([booking for booking in bookings if booking.get("booking_status") == "pending"]),
            "soft_lock": len([booking for booking in bookings if booking.get("booking_status") == "soft_lock"]),
            "confirmed": len([booking for booking in bookings if booking.get("booking_status") == "confirmed"]),
            "completed": len([booking for booking in bookings if booking.get("booking_status") == "completed"]),
            "cancelled": len([booking for booking in bookings if booking.get("booking_status") == "cancelled"]),
            "upcoming": len([booking for booking in bookings if booking.get("booking_status") in {"confirmed", "soft_lock"} and _booking_date(booking.get("check_in_date")) and _booking_date(booking.get("check_in_date")) >= today]),
            "checked_in": len([booking for booking in bookings if booking.get("booking_status") == "confirmed" and _booking_date(booking.get("check_in_date")) and _booking_date(booking.get("check_in_date")) <= today and _booking_date(booking.get("check_out_date")) and _booking_date(booking.get("check_out_date")) >= today]),
            "checked_out": len([booking for booking in bookings if booking.get("booking_status") in {"confirmed", "completed"} and _booking_date(booking.get("check_out_date")) and _booking_date(booking.get("check_out_date")) < today]),
            "revenue": round(sum(float(booking.get("total_amount") or 0) for booking in revenue_bookings), 2),
            "occupancy": round((len(revenue_bookings) / max(len(property_ids) * 30, 1)) * 100, 1),
        }
        return {"bookings": bookings, "summary": summary, "total": len(bookings)}

    except Exception as e:
        logger.error(f"Error fetching RM bookings: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to fetch bookings")


@router.get("/bookings/{booking_id}/details")
async def get_rm_booking_details(
    booking_id: str,
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Get detailed booking view for RM scoped booking management."""
    try:
        _, _, _, booking_query = await _get_rm_booking_query(db, current_user["user_id"])
        booking = await db.bookings.find_one({"$and": [booking_query, {"booking_id": booking_id}]}, {"_id": 0})
        if not booking:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Booking not found or not assigned to you")

        prop = await db.properties.find_one({"property_id": booking.get("property_id")}, {"_id": 0})
        host = await db.users.find_one({"user_id": booking.get("host_id") or (prop or {}).get("owner_id")}, {"_id": 0, "password_hash": 0})
        broker = await db.users.find_one({"user_id": booking.get("broker_id") or (prop or {}).get("broker_id")}, {"_id": 0, "password_hash": 0}) if booking.get("broker_id") or (prop or {}).get("broker_id") else None
        guest = await db.users.find_one({"user_id": booking.get("guest_id")}, {"_id": 0, "password_hash": 0})
        commissions = await db.commissions.find({"booking_id": booking_id}, {"_id": 0}).sort("created_at", -1).to_list(length=50)
        audit_logs = await db.audit_logs.find({"record_id": booking_id}, {"_id": 0}).sort("created_at", -1).to_list(length=50)
        timeline = [
            {"label": "Booking Created", "status": "completed", "created_at": booking.get("created_at")},
            {"label": "Payment", "status": booking.get("payment_status") or "pending", "created_at": booking.get("confirmed_at")},
            {"label": "Check-in", "status": "scheduled", "created_at": booking.get("check_in_date")},
            {"label": "Check-out", "status": "scheduled", "created_at": booking.get("check_out_date")},
        ]
        return {
            "booking": booking,
            "property": prop,
            "host": host,
            "broker": broker,
            "guest": guest,
            "commissions": commissions,
            "audit_logs": audit_logs,
            "timeline": timeline,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching RM booking details: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to fetch booking details")

# ========== TASKS, ESCALATIONS AND NOTIFICATIONS ==========

def _parse_dt(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            return None
    return None


def _rm_task_sla_status(created_at, sla_hours=24):
    created = _parse_dt(created_at) or datetime.now(timezone.utc)
    if created.tzinfo is None:
        created = created.replace(tzinfo=timezone.utc)
    age_hours = (datetime.now(timezone.utc) - created).total_seconds() / 3600
    if age_hours >= sla_hours * 2:
        return "escalated", round(age_hours, 1)
    if age_hours >= sla_hours:
        return "breached", round(age_hours, 1)
    if age_hours >= max(0, sla_hours - 4):
        return "at_risk", round(age_hours, 1)
    return "within_sla", round(age_hours, 1)


@router.get("/tasks")
async def get_rm_tasks(
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Get RM operational tasks, escalation watchlist and notification feed."""
    try:
        rm_id = current_user["user_id"]
        rm_identifiers = await _get_rm_identifiers(db, current_user)
        broker_ids, host_ids, property_query = await _get_rm_property_query(db, rm_id)
        properties = await db.properties.find(property_query, {"_id": 0, "property_id": 1, "title": 1, "status": 1, "created_at": 1, "updated_at": 1}).to_list(length=5000)
        property_ids = [prop["property_id"] for prop in properties if prop.get("property_id")]
        tasks = []

        pending_rm_reviews = await db.property_verifications.find(
            {
                **_verification_scope_query(broker_ids, property_ids, rm_identifiers),
                "status": VerificationStatus.COMPLETED.value,
                "rm_reviewed": False,
            },
            {"_id": 0}
        ).sort("completed_at", -1).to_list(length=200)
        for item in pending_rm_reviews:
            sla_status, age_hours = _rm_task_sla_status(item.get("completed_at") or item.get("created_at"), 48)
            tasks.append({
                "task_id": item.get("verification_id"),
                "type": "rm_verification",
                "title": "Pending RM Verification",
                "entity_id": item.get("property_id"),
                "priority": "high" if sla_status in {"at_risk", "breached", "escalated"} else "normal",
                "status": "open",
                "sla_status": sla_status,
                "age_hours": age_hours,
                "due_label": "48h RM review SLA",
                "created_at": item.get("completed_at") or item.get("created_at"),
            })

        missing_doc_hosts = await db.users.find(
            {"user_id": {"$in": host_ids}, "role": "host", "kyc_status": {"$ne": "approved"}},
            {"_id": 0, "user_id": 1, "full_name": 1, "kyc_status": 1, "created_at": 1, "broker_id": 1}
        ).to_list(length=200)
        for host in missing_doc_hosts:
            sla_status, age_hours = _rm_task_sla_status(host.get("created_at"), 72)
            tasks.append({
                "task_id": f"host_docs_{host.get('user_id')}",
                "type": "host_documents",
                "title": f"Pending Host Documents - {host.get('full_name') or host.get('user_id')}",
                "entity_id": host.get("user_id"),
                "priority": "high" if sla_status in {"breached", "escalated"} else "normal",
                "status": host.get("kyc_status") or "pending",
                "sla_status": sla_status,
                "age_hours": age_hours,
                "due_label": "72h host document SLA",
                "created_at": host.get("created_at"),
            })

        for prop in properties:
            if prop.get("status") in {"draft", "pending_verification", "under_review", "rejected"}:
                sla_status, age_hours = _rm_task_sla_status(prop.get("updated_at") or prop.get("created_at"), 48)
                tasks.append({
                    "task_id": f"property_review_{prop.get('property_id')}",
                    "type": "property_review",
                    "title": f"Pending Property Review - {prop.get('title') or prop.get('property_id')}",
                    "entity_id": prop.get("property_id"),
                    "priority": "high" if sla_status in {"breached", "escalated"} else "normal",
                    "status": prop.get("status") or "draft",
                    "sla_status": sla_status,
                    "age_hours": age_hours,
                    "due_label": "48h property review SLA",
                    "created_at": prop.get("updated_at") or prop.get("created_at"),
                })

        support_tickets = await db.support_tickets.find(
            {
                "status": {"$nin": ["resolved", "closed"]},
                "$or": [
                    *_field_matches_identifiers("user_rm_id", rm_identifiers),
                    *_field_matches_identifiers("rm_id", rm_identifiers),
                    *_field_matches_identifiers("assigned_rm_id", rm_identifiers),
                    {"user_broker_id": {"$in": broker_ids}},
                    {"broker_id": {"$in": broker_ids}},
                    {"user_id": {"$in": host_ids}},
                    {"host_id": {"$in": host_ids}},
                    {"property_id": {"$in": property_ids}},
                ],
            },
            {"_id": 0}
        ).sort("created_at", -1).to_list(length=100)
        for ticket in support_tickets:
            sla_status, age_hours = _rm_task_sla_status(ticket.get("sla_due_at") or ticket.get("created_at"), 24)
            tasks.append({
                "task_id": ticket.get("ticket_id"),
                "type": "support_ticket",
                "title": ticket.get("subject") or "Support ticket",
                "entity_id": ticket.get("ticket_id"),
                "priority": ticket.get("priority") or "normal",
                "status": ticket.get("status") or "open",
                "sla_status": sla_status,
                "age_hours": age_hours,
                "due_label": "24h support response SLA",
                "created_at": ticket.get("created_at"),
            })

        escalation_instances = await db.escalation_instances.find(
            {
                "status": {"$nin": ["resolved", "closed"]},
                "$or": [
                    *_field_matches_identifiers("assigned_rm_id", rm_identifiers),
                    *_field_matches_identifiers("rm_id", rm_identifiers),
                    {"broker_id": {"$in": broker_ids}},
                    {"host_id": {"$in": host_ids}},
                    {"property_id": {"$in": property_ids}},
                ],
            },
            {"_id": 0}
        ).sort("created_at", -1).to_list(length=100)

        notifications = await db.notifications.find(
            {
                "$or": [
                    *_field_matches_identifiers("user_id", rm_identifiers),
                    *_field_matches_identifiers("data.rm_id", rm_identifiers),
                    {"data.broker_id": {"$in": broker_ids}},
                    {"data.host_id": {"$in": host_ids}},
                    {"data.property_id": {"$in": property_ids}},
                ]
            },
            {"_id": 0}
        ).sort("created_at", -1).limit(50).to_list(length=50)

        tasks = sorted(tasks, key=lambda item: item.get("age_hours", 0), reverse=True)[:300]
        escalations = [task for task in tasks if task.get("sla_status") in {"at_risk", "breached", "escalated"}]
        escalations.extend([{
            "task_id": item.get("escalation_id") or item.get("id"),
            "type": item.get("type") or "escalation",
            "title": item.get("title") or item.get("reason") or "Escalation",
            "entity_id": item.get("property_id") or item.get("host_id") or item.get("broker_id"),
            "priority": item.get("priority") or "high",
            "status": item.get("status") or "open",
            "sla_status": item.get("sla_status") or "escalated",
            "age_hours": _rm_task_sla_status(item.get("created_at"), 1)[1],
            "due_label": "Escalation matrix",
            "created_at": item.get("created_at"),
        } for item in escalation_instances])

        return {
            "tasks": tasks,
            "escalations": escalations[:100],
            "notifications": notifications,
            "summary": {
                "open_tasks": len(tasks),
                "critical_tasks": len([task for task in tasks if task.get("priority") in {"high", "urgent"}]),
                "overdue_tasks": len([task for task in tasks if task.get("sla_status") in {"breached", "escalated"}]),
                "sla_breaches": len([task for task in tasks if task.get("sla_status") in {"breached", "escalated"}]),
                "pending_approvals": len(pending_rm_reviews),
                "notifications": len(notifications),
                "escalations": len(escalations),
            }
        }

    except Exception as e:
        logger.error(f"Error fetching RM tasks: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to fetch tasks")

# ========== REPORTS ==========

@router.get("/reports/rm-analytics-overview")
async def get_rm_analytics_overview(
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Get RM scoped analytics across brokers, hosts, properties, bookings and SLA."""
    try:
        rm_id = current_user["user_id"]
        rm_identifiers = await _get_rm_identifiers(db, current_user)
        broker_ids, host_ids, property_query = await _get_rm_property_query(db, rm_id)
        properties = await db.properties.find(property_query, {"_id": 0}).to_list(length=5000)
        property_ids = [prop["property_id"] for prop in properties if prop.get("property_id")]
        _, _, _, booking_query = await _get_rm_booking_query(db, rm_id)
        bookings = await db.bookings.find(
            booking_query,
            {"_id": 0}
        ).to_list(length=10000)
        commissions = await db.commissions.find({"broker_id": {"$in": broker_ids}}, {"_id": 0}).to_list(length=5000)
        verifications = await db.property_verifications.find(
            _verification_scope_query(broker_ids, property_ids, rm_identifiers),
            {"_id": 0}
        ).to_list(length=5000)

        paid_bookings = [
            booking for booking in bookings
            if booking.get("payment_status") in {"paid", "partially_paid"} or booking.get("booking_status") in {"confirmed", "completed"}
        ]
        revenue = round(sum(float(booking.get("total_amount") or 0) for booking in paid_bookings), 2)
        commission_total = round(sum(float(item.get("commission_amount") or 0) for item in commissions), 2)
        confirmed_bookings = len([booking for booking in bookings if booking.get("booking_status") in {"confirmed", "completed"}])
        conversion_rate = round((confirmed_bookings / max(len(bookings), 1)) * 100, 1)
        live_properties = len([prop for prop in properties if prop.get("status") == "live"])
        pending_properties = len([prop for prop in properties if prop.get("status") in {"draft", "pending", "pending_verification", "under_review"}])
        rejected_properties = len([prop for prop in properties if prop.get("status") == "rejected"])
        approved_verifications = len([item for item in verifications if item.get("rm_reviewed") or item.get("admin_approved")])
        verification_rate = round((approved_verifications / max(len(verifications), 1)) * 100, 1)
        sla_breaches = len([
            item for item in verifications
            if not item.get("rm_reviewed") and _rm_task_sla_status(item.get("completed_at") or item.get("created_at"), 48)[0] in {"breached", "escalated"}
        ])

        broker_rows = []
        brokers = await db.users.find({"user_id": {"$in": broker_ids}}, {"_id": 0, "password_hash": 0}).to_list(length=1000)
        for broker in brokers:
            broker_id = broker.get("user_id")
            broker_properties = [prop for prop in properties if prop.get("broker_id") == broker_id]
            broker_property_ids = [prop["property_id"] for prop in broker_properties if prop.get("property_id")]
            broker_bookings = [
                booking for booking in bookings
                if booking.get("broker_id") == broker_id or booking.get("property_id") in broker_property_ids
            ]
            broker_paid = [
                booking for booking in broker_bookings
                if booking.get("payment_status") in {"paid", "partially_paid"} or booking.get("booking_status") in {"confirmed", "completed"}
            ]
            broker_rows.append({
                "broker_id": broker_id,
                "broker_name": broker.get("full_name") or broker_id,
                "lg_code": broker.get("lg_code") or "N/A",
                "hosts": len([host_id for host_id in host_ids if any(prop.get("owner_id") == host_id and prop.get("broker_id") == broker_id for prop in properties)]),
                "properties": len(broker_properties),
                "live_properties": len([prop for prop in broker_properties if prop.get("status") == "live"]),
                "bookings": len(broker_bookings),
                "revenue": round(sum(float(booking.get("total_amount") or 0) for booking in broker_paid), 2),
                "commission": round(sum(float(item.get("commission_amount") or 0) for item in commissions if item.get("broker_id") == broker_id), 2),
            })

        host_rows = []
        hosts = await db.users.find({"user_id": {"$in": host_ids}}, {"_id": 0, "password_hash": 0}).to_list(length=1000)
        for host in hosts:
            host_id = host.get("user_id")
            host_properties = [prop for prop in properties if prop.get("owner_id") == host_id]
            host_property_ids = [prop["property_id"] for prop in host_properties if prop.get("property_id")]
            host_bookings = [
                booking for booking in bookings
                if booking.get("host_id") == host_id or booking.get("property_id") in host_property_ids
            ]
            host_paid = [
                booking for booking in host_bookings
                if booking.get("payment_status") in {"paid", "partially_paid"} or booking.get("booking_status") in {"confirmed", "completed"}
            ]
            host_rows.append({
                "host_id": host_id,
                "host_name": host.get("full_name") or host_id,
                "kyc_status": host.get("kyc_status") or "pending",
                "properties": len(host_properties),
                "live_properties": len([prop for prop in host_properties if prop.get("status") == "live"]),
                "bookings": len(host_bookings),
                "revenue": round(sum(float(booking.get("total_amount") or 0) for booking in host_paid), 2),
            })

        property_rows = [{
            "property_id": prop.get("property_id"),
            "title": prop.get("title") or prop.get("property_id"),
            "city": prop.get("city") or "N/A",
            "status": prop.get("status") or "draft",
            "broker_id": prop.get("broker_id"),
            "host_id": prop.get("owner_id"),
            "bookings": len([booking for booking in bookings if booking.get("property_id") == prop.get("property_id")]),
            "revenue": round(sum(float(booking.get("total_amount") or 0) for booking in paid_bookings if booking.get("property_id") == prop.get("property_id")), 2),
        } for prop in properties]

        return {
            "report_type": "rm_analytics_overview",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "summary": {
                "brokers": len(broker_ids),
                "hosts": len(host_ids),
                "properties": len(properties),
                "live_properties": live_properties,
                "pending_properties": pending_properties,
                "rejected_properties": rejected_properties,
                "bookings": len(bookings),
                "confirmed_bookings": confirmed_bookings,
                "revenue": revenue,
                "commission": commission_total,
                "conversion_rate": conversion_rate,
                "verification_rate": verification_rate,
                "sla_breaches": sla_breaches,
            },
            "brokers": sorted(broker_rows, key=lambda item: item["revenue"], reverse=True),
            "hosts": sorted(host_rows, key=lambda item: item["revenue"], reverse=True),
            "properties": sorted(property_rows, key=lambda item: item["revenue"], reverse=True)[:300],
        }

    except Exception as e:
        logger.error(f"Error fetching RM analytics overview: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to fetch analytics overview")


@router.get("/reports/rm-analytics-overview/export-csv")
async def export_rm_analytics_overview_csv(
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Export RM analytics property performance as CSV."""
    data = await get_rm_analytics_overview(current_user=current_user, db=db)
    output = io.StringIO()
    fieldnames = ["Property ID", "Title", "City", "Status", "Broker ID", "Host ID", "Bookings", "Revenue"]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for prop in data.get("properties", []):
        writer.writerow({
            "Property ID": prop.get("property_id"),
            "Title": prop.get("title"),
            "City": prop.get("city"),
            "Status": prop.get("status"),
            "Broker ID": prop.get("broker_id"),
            "Host ID": prop.get("host_id"),
            "Bookings": prop.get("bookings"),
            "Revenue": prop.get("revenue"),
        })
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=rm_analytics_{datetime.now(timezone.utc).strftime('%Y%m%d')}.csv"}
    )


@router.get("/audit-activity")
async def get_rm_audit_activity(
    module: Optional[str] = None,
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Get RM-scoped audit activity across assigned brokers, hosts, properties and bookings."""
    try:
        rm_id = current_user["user_id"]
        rm_identifiers = await _get_rm_identifiers(db, current_user)
        broker_ids, host_ids, property_query = await _get_rm_property_query(db, rm_id)
        properties = await db.properties.find(property_query, {"_id": 0, "property_id": 1}).to_list(length=5000)
        property_ids = [prop["property_id"] for prop in properties if prop.get("property_id")]
        _, _, _, booking_query = await _get_rm_booking_query(db, rm_id)
        bookings = await db.bookings.find(
            booking_query,
            {"_id": 0, "booking_id": 1}
        ).to_list(length=5000)
        booking_ids = [booking["booking_id"] for booking in bookings if booking.get("booking_id")]
        scoped_record_ids = list({rm_id, *rm_identifiers, *broker_ids, *host_ids, *property_ids, *booking_ids})
        audit_query = {
            "$or": [
                *_field_matches_identifiers("user_id", rm_identifiers),
                {"record_id": {"$in": scoped_record_ids}},
                *_field_matches_identifiers("new_value.rm_id", rm_identifiers),
                *_field_matches_identifiers("old_value.rm_id", rm_identifiers),
            ]
        }
        if module and module != "all":
            audit_query = {"$and": [audit_query, {"module": module}]}

        audit_logs = await db.audit_logs.find(audit_query, {"_id": 0}).sort("created_at", -1).limit(300).to_list(length=300)
        failed_events = len([item for item in audit_logs if item.get("status") not in {None, "success"}])
        approval_events = len([item for item in audit_logs if "approve" in (item.get("action") or "").lower() or "reject" in (item.get("action") or "").lower()])
        modules = sorted(list({item.get("module") or "general" for item in audit_logs}))

        return {
            "audit_logs": audit_logs,
            "summary": {
                "total_events": len(audit_logs),
                "approval_events": approval_events,
                "failed_events": failed_events,
                "modules": len(modules),
                "scoped_records": len(scoped_record_ids),
            },
            "filters": {
                "module": module or "all",
                "available_modules": modules,
            },
        }

    except Exception as e:
        logger.error(f"Error fetching RM audit activity: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to fetch audit activity")


@router.get("/reports/properties-not-booked")
async def get_properties_not_booked_report(
    broker_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Get report of active properties with zero bookings."""
    try:
        # Get my brokers
        my_broker_ids, my_host_ids, scoped_property_query = await _get_rm_property_query(db, current_user["user_id"])

        # Build property query
        property_query = {"$and": [scoped_property_query, {"status": "live"}]}
        if broker_id:
            if broker_id not in my_broker_ids:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Broker is not assigned to you"
                )
            property_query = {"$and": [scoped_property_query, {"status": "live"}, {"broker_id": broker_id}]}
        
        # Get all matching live properties
        property_cursor = db.properties.find(property_query, {"_id": 0})
        properties = await property_cursor.to_list(length=500)
        
        # For each property, check if it has bookings
        properties_not_booked = []
        
        for property in properties:
            booking_count = await db.bookings.count_documents({
                "property_id": property["property_id"],
                "booking_status": "confirmed"
            })
            
            if booking_count == 0:
                # Get broker details
                if property.get("broker_id"):
                    broker = await db.users.find_one(
                        {"user_id": property["broker_id"]},
                        {"_id": 0, "full_name": 1, "lg_code": 1}
                    )
                    property["broker_name"] = broker.get("full_name") if broker else "N/A"
                    property["broker_lg_code"] = broker.get("lg_code") if broker else "N/A"
                
                properties_not_booked.append(property)
        
        return {
            "report_type": "properties_not_booked",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "filters": {
                "broker_id": broker_id,
                "start_date": start_date,
                "end_date": end_date
            },
            "properties": properties_not_booked,
            "total": len(properties_not_booked)
        }
    
    except Exception as e:
        logger.error(f"Error generating report: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate report"
        )

@router.get("/reports/properties-not-booked/export-csv")
async def export_properties_not_booked_csv(
    broker_id: Optional[str] = None,
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Export properties not booked report as CSV."""
    try:
        # Get my brokers
        my_broker_ids, my_host_ids, scoped_property_query = await _get_rm_property_query(db, current_user["user_id"])

        # Get report data
        property_query = {"$and": [scoped_property_query, {"status": "live"}]}
        if broker_id:
            if broker_id not in my_broker_ids:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Broker is not assigned to you"
                )
            property_query = {"$and": [scoped_property_query, {"status": "live"}, {"broker_id": broker_id}]}
        
        property_cursor = db.properties.find(property_query, {"_id": 0})
        properties = await property_cursor.to_list(length=500)
        
        properties_not_booked = []
        
        for property in properties:
            booking_count = await db.bookings.count_documents({
                "property_id": property["property_id"],
                "booking_status": "confirmed"
            })
            
            if booking_count == 0:
                if property.get("broker_id"):
                    broker = await db.users.find_one(
                        {"user_id": property["broker_id"]},
                        {"_id": 0, "full_name": 1, "lg_code": 1}
                    )
                    broker_name = broker.get("full_name") if broker else "N/A"
                    broker_lg_code = broker.get("lg_code") if broker else "N/A"
                else:
                    broker_name = "N/A"
                    broker_lg_code = "N/A"
                
                properties_not_booked.append({
                    "Property ID": property["property_id"],
                    "Title": property["title"],
                    "City": property["city"],
                    "BHK Type": property["bhk_type"],
                    "Category": property["category"],
                    "Price per Night": property.get("price_per_night", 0),
                    "Broker Name": broker_name,
                    "Broker LG Code": broker_lg_code,
                    "Created At": property["created_at"].isoformat() if isinstance(property["created_at"], datetime) else property["created_at"]
                })
        
        # Create CSV
        output = io.StringIO()
        if properties_not_booked:
            fieldnames = properties_not_booked[0].keys()
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(properties_not_booked)
        
        # Return as streaming response
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=properties_not_booked_{datetime.now(timezone.utc).strftime('%Y%m%d')}.csv"
            }
        )
    
    except Exception as e:
        logger.error(f"Error exporting CSV: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to export CSV"
        )

@router.get("/reports/broker-portfolio-summary")
async def get_broker_portfolio_summary(
    current_user: dict = Depends(require_employee),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Get summary report of all brokers' portfolios."""
    try:
        broker_ids, _ = await _get_rm_scope(db, current_user)
        broker_cursor = db.users.find({"role": "broker", "user_id": {"$in": broker_ids}}, {"_id": 0})
        brokers = await broker_cursor.to_list(length=200)
        
        summary = []
        
        for broker in brokers:
            broker_id = broker["user_id"]
            
            total_properties = await db.properties.count_documents({"broker_id": broker_id})
            live_properties = await db.properties.count_documents({"broker_id": broker_id, "status": "live"})
            pending_verification = await db.property_verifications.count_documents({
                "broker_id": broker_id,
                "status": {"$in": ["pending", "in_progress"]}
            })
            
            summary.append({
                "broker_name": broker["full_name"],
                "lg_code": broker.get("lg_code", "N/A"),
                "total_properties": total_properties,
                "live_properties": live_properties,
                "pending_verification": pending_verification
            })
        
        return {
            "report_type": "broker_portfolio_summary",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "brokers": summary,
            "total": len(summary)
        }
    
    except Exception as e:
        logger.error(f"Error generating broker portfolio summary: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate report"
        )
